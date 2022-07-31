from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side

data = {
    'Fred' : {
        'Week 1':100 ,
        'Week 2':100 ,
        'Week 3':106,
        'Week 4':105,
        'Week 5':106
    },
    'Dan': {
        'Week 1':105 ,
        'Week 2':0 ,
        'Week 3':105,
        'Week 4':106,
        'Week 5':105
    },
    'Grace': {
        'Week 1':100,
        'Week 2':83 ,
        'Week 3':106,
        'Week 4':105,
        'Week 5':106
    },
    'Ruth' :{
        'Week 1':102,
        'Week 2':100 ,
        'Week 3':106,
        'Week 4':105,
        'Week 5':106
    },
    'Kelvin' :{
        'Week 1':100,
        'Week 2':117,
        'Week 3':106,
        'Week 4':105,
        'Week 5':106        
    },
    'Rodney' :{
        'Week 1':105,
        'Week 2':100 ,
        'Week 3':106,
        'Week 4':0,
        'Week 5':106
    },
    'Eunice' :{
        'Week 1':105,
        'Week 2':101 ,
        'Week 3':106,
        'Week 4':105,
        'Week 5':106
    }
}
wb = Workbook()
ws = wb.active

# naming the sheet
ws.title = 'WeeklyOutput'

# changing the background colour of the worksheet.
ws.sheet_properties.tabColor = "0072BA"

headings =['Name'] + list(data['Fred'].keys())
ws.append(headings)

for person in data:
    output = list(data[person].values())
    ws.append([person] + output)


# styling the cells
for col in range (1,7):
    ws[get_column_letter(col) + '1'].font = Font(
        name = 'Calibri',size = 10, bold =True, italic=True, vertAlign= 'baseline' , underline= 'none',strike= False, color='00000000'
    )

for col in ws['A']:
    col.font = Font(name='Calibri', size=10, bold=True,color='00800000')

for row in ws[9]:
    row.font = Font (name = 'Calibri', size=10,italic=True,color='000000FF')

cell1 = ws['C3']
cell1.font = Font (name = 'Calibri', size=10,bold=True,color='00FF0000')

cell2 = ws['E7']
cell2.font = Font (name = 'Calibri', size=10,bold=True,color='00FF0000')


# inserting a column between 'Name' and 'Week 1'
ws.insert_cols(2)

# naming the column
ws['B1'] = 'Task'
cell3 = ws['B1']
cell3.font = Font (name = 'Calibri', size=10,bold=True,color='080808')

ws['B2'] = 'Samsung Phones'
ws['B3'] = 'Nokia Phones'
ws['B4'] = '32" Frames'
ws['B5'] = 'IPQC'
ws['B5'] = 'Samsung Phones'
ws['B6'] = 'Nokia Phones'
ws['B7'] = 'Speaker case'
ws['B8'] = 'Accessories'

# average performance
for col in range(2, len(data['Fred']) + 3):
    char = get_column_letter(col)
    ws[char + '9'] = f"=SUM({char + '2'}:{char + '8'})/{len(data)}"

ws['A9'] = 'Average'
ws['B9'] = ' '

# setting boarders
def set_border(ws, cell_range):
    thin = Side(border_style="thin", color="080808")
    for row in ws[cell_range]:
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

set_border(ws, 'A1:G9') 

ws.auto_filter.ref = 'A1:G1'


wb.save('Team_Performance.xlsx')

